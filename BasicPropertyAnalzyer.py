#!/usr/bin/python
# -*- coding: iso-8859-15 -*-
import sys
# IO package für Pfadnamen
stflist = ['C:\\Program Files\\Stimfit 0.14\\wx-3.0-msw', 
           'C:\\Program Files\\Stimfit 0.14\\wx-3.0-msw', 
           'C:\\Program Files\\Stimfit 0.14', 
           'C:\\Program Files\\Stimfit 0.14\\stf-site-packages', 
           'C:\\WINDOWS\\SYSTEM32\\python27.zip', 
           'C:\\Users\\c-sch_000\\Anaconda\\Lib', 
           'C:\\Users\\c-sch_000\\Anaconda\\DLLs', 
           'C:\\Python27\\Lib', 
           'C:\\Python27\\DLLs', 
           'C:\\Python27\\Lib\\lib-tk', 
           'C:\\Program Files\\Stimfit 0.14', 
           'C:\\Python27', 
           'C:\\Python27\\lib\\site-packages']
sys.path = list(set(sys.path + stflist))
from Tkinter import *
import tkMessageBox as box
import tkFileDialog 
import stfio
import tkSimpleDialog 

# Imports for Export to Excel
import openpyxl
from openpyxl import Workbook
from openpyxl.compat import range
from openpyxl.cell import get_column_letter
from openpyxl.styles import Font, Fill

# Imports for Peak Detection and Calculations
import numpy as np
from math import pi, log
import pylab
from scipy import fft, ifft, stats
from scipy.optimize import curve_fit

# Imports for Matplotlib
import matplotlib.pyplot as plt

# Import for directory settings

import os
from time import strftime

rec = stfio.read("14122000.abf")
last_value = 0
complete_dataset = np.array([0,0])
rig = 0
abffile = "None"
iAP_file = "None"
input_resistance = 0
capacitance = 1000
openDirectory = "C:\\"
saveDirectory = "C:\\temp\\"

#Create Excel file
wb = Workbook()
#dest_filename = saveDirectory + "\\"  + strftime("%Y-%m-%d_%H-%M-%S") + ".xlsx"
dest_filename = ""
dest_directory = ""

ws1 = wb.active
ws1.title = "Test"

# Filling of Column A with Descriptions
ws1['A5'] = "Filename of the first CC recording"
ws1['A6'] = "rig"
ws1['A7'] = "Vmembrane (mV)"
ws1['A8'] = "1-5 slope input res (Mohms)"
ws1['A9'] = "Capacitance (pf)"  
ws1['A10'] = "Series resistance (Mohm)"
ws1['A11'] = "Voltage Offset (mV)"
ws1['A12'] = "sAP"
ws1['A13'] = "n (1)"
ws1['A14'] = "a (2)"
ws1['A15'] = "y (3)"
ws1['A16'] = "sAP code"
ws1['A17'] = "sAP description (burst | single)"
ws1['A18'] = "iAP"
ws1['A19'] = "n (1)"
ws1['A20'] = "a (2)"
ws1['A21'] = "ys (3)"
ws1['A22'] = "at (4)"
ws1['A23'] = "yt (5)"
ws1['A24'] = "iAP code"


ws1['A27'] = "filename"
ws1['A28'] = "sweep number"

ws1['A30'] = "Overshoot (mv)"
ws1['A31'] = "Afterhyperpolarization (mv)"
ws1['A32'] = "(calculated) Spike Height (mV)"

ws1['A81'] = "I Na max (pA/pF)"
ws1['A82'] = "I K max (pA/pF)"

ws1['A84'] = "Na activation (pA)"

ws1['A122'] = "Na activation (pA/pF)"

ws1['A160'] = "U = R*I, (120-66.68)"
ws1['A161'] = "Na activation driving force corrected (66.68 mV)"

ws1['A201'] = "Na activation, normalized conductance G/Gmax"

ws1['A241'] = "Na inactivation (pA)"

ws1['A280'] = "Na inactivation conductance (driving force corrected (66.68 mV))"

ws1['A320'] = "Na inactivation, normalized conductance G/Gmax"

ws1['A359'] = "K currents (pA) last 50 ms averaged"

ws1['B24'] = "=SUM(B19:B23)"


def voltageStepInserter(first_column, last_column, voltage_begin, step):
    m = 0
    for k in range(first_column,last_column+1):
        coordinate = "A" + str(k)
        ws1[coordinate] = voltage_begin-(m*-step)
        m = m + 1   
# instert steps for current step protocol
voltageStepInserter(37, 56, -10, 10)

#wb.save(filename = dest_filename)


def _datacheck_peakdetect(x_axis, y_axis):
    if x_axis is None:
        x_axis = range(len(y_axis))
    
    if len(y_axis) != len(x_axis):
        raise (ValueError, 
                'Input vectors y_axis and x_axis must have same length')
    
    #needs to be a numpy array
    y_axis = np.array(y_axis)
    x_axis = np.array(x_axis)
    return x_axis, y_axis


def peakdetect(y_axis, direction, x_axis = None, lookahead = 100, delta=0):
    global min_peak
    """
    Converted from/based on a MATLAB script at: 
    http://billauer.co.il/peakdet.html
    found at https://gist.github.com/sixtenbe/1178136
    
    function for detecting local maximas and minmias in a signal.
    Discovers peaks by searching for values which are surrounded by lower
    or larger values for maximas and minimas respectively
    
    keyword arguments:
    y_axis -- A list containg the signal over which to find peaks
    x_axis -- (optional) A x-axis whose values correspond to the y_axis list
        and is used in the return to specify the postion of the peaks. If
        omitted an index of the y_axis is used. (default: None)
    lookahead -- (optional) distance to look ahead from a peak candidate to
        determine if it is the actual peak (default: 200) 
        '(sample / period) / f' where '4 >= f >= 1.25' might be a good value
    delta -- (optional) this specifies a minimum difference between a peak and
        the following points, before a peak may be considered a peak. Useful
        to hinder the function from picking up false peaks towards to end of
        the signal. To work well delta should be set to delta >= RMSnoise * 5.
        (default: 0)
            delta function causes a 20% decrease in speed, when omitted
            Correctly used it can double the speed of the function
    
    return -- two lists [max_peaks, min_peaks] containing the positive and
        negative peaks respectively. Each cell of the lists contains a tupple
        of: (position, peak_value) 
        to get the average peak value do: np.mean(max_peaks, 0)[1] on the
        results to unpack one of the lists into x, y coordinates do: 
        x, y = zip(*tab)
    """
    max_peaks = []
    min_peaks = []
    min_peak = 10000000
    dump = []   #Used to pop the first hit which almost always is false
       
    # check input data
    x_axis, y_axis = _datacheck_peakdetect(x_axis, y_axis)
    # store data length for later use
    length = len(y_axis)
    
    
    #perform some checks
    if lookahead < 1:
        raise ValueError, "Lookahead must be '1' or above in value"
    if not (np.isscalar(delta) and delta >= 0):
        raise ValueError, "delta must be a positive number"
    
    #maxima and minima candidates are temporarily stored in
    #mx and mn respectively
    mn, mx = np.Inf, -np.Inf
    
    #Only detect peak if there is 'lookahead' amount of points after it
    for index, (x, y) in enumerate(zip(x_axis[:-lookahead], 
                                        y_axis[:-lookahead])):
        if y > mx:
            mx = y
            mxpos = x
        if y < mn:
            mn = y
            mnpos = x
        
        ####look for max####
        if y < mx-delta and mx != np.Inf:
            #Maxima peak candidate found
            #look ahead in signal to ensure that this is a peak and not jitter
            if y_axis[index:index+lookahead].max() < mx:
                if y_axis[mxpos] > 0: 
                    max_peaks.append([mxpos, mx])
                    #dump.append(True)
                #set algorithm to only find minima now
                mx = np.Inf
                mn = np.Inf
                if index+lookahead >= length:
                    #end is within lookahead no more peaks can be found
                    break
                continue
            #else:  #slows shit down this does
            #    mx = ahead
            #    mxpos = x_axis[np.where(y_axis[index:index+lookahead]==mx)]
        
        ####look for min####
        
        if y > mn+delta and mn != -np.Inf:
            #Minima peak candidate found 
            #look ahead in signal to ensure that this is a peak and not jitter
            if y_axis[index:index+lookahead].min() > mn:
                #print y_axis[mnpos] 
                min_peaks.append([mnpos, mn])
                if y_axis[mnpos] < min_peak:
                    min_peak = y_axis[mnpos]
                    
                #dump.append(False)
                #set algorithm to only find maxima now
                mn = -np.Inf
                mx = -np.Inf
                if index+lookahead >= length:
                    #end is within lookahead no more peaks can be found
                    break
            #else:  #slows shit down this does
            #    mn = ahead
            #    mnpos = x_axis[np.where(y_axis[index:index+lookahead]==mn)]
    if direction == "positive":
        return max_peaks 
        
    if direction == "negative":
        return min_peak
       



class spontActivity():
    def attempted_action_potential_found(complete_section):
        #Sucht nach versuchten Aktionspotentialen (Peak hoeher als -20 mV) im gesamten Trace
        return complete_dataset[complete_dataset.argmax()] > -10


    def action_potential_found(complete_section):
        #Sucht nach Aktionspotentialen im gesamten Trace
        return complete_dataset[complete_dataset.argmax()] > 0

        
    def mean_first_10s(self, complete_section):     
        global last_value, mean_array
        #Calculates the mean of the first 10 s
        first_value = 0
        # 10000 ms period divided by sampling rate
        # for example: 10.000/0.2 (0.2 ms interval) = 50.000 data points
        # for example 2: 10.000/0.05 (0.05 ms interval) = 50.000 data points
        last_value = 10000/rec.dt

        mean_array = complete_section[first_value:last_value]

        return mean_array.mean()
        
class inducedActivity():
    def calculateInputResistance(self, first_mean, second_mean):
        global input_resistance
        '''
        Receives means of two current steps in mV
        Assumes Delta of 10 pA
        Returns the input resistance in MOhm, calculated with Ohm's law U = R * I
        '''
        #transform pA in A
        deltaI = 1E-11
        
        #DeltaU and transform mV in V
        deltaU = (second_mean-first_mean)/1000
        
        # U = R * I => R = U/I
        resistance = deltaU/deltaI
        
        #tranform Ohm in MOhm
        input_resistance = resistance/1000000
        
        return input_resistance

class Example(Frame):
  
    def __init__(self, parent):
        Frame.__init__(self, parent)   
         
        self.parent = parent        
        self.initUI()
        
    def initUI(self):
      
        self.parent.title("Basic property analyzer for the PJK lab")
        self.pack(fill=BOTH, expand=1)
        menubar = Menu(self.parent)
        self.parent.config(menu=menubar)
        
        fileMenu = Menu(menubar, tearoff=0)
        fileMenu.add_command(label="Open file gap free recording...", command=self.onOpenGapFree)
        fileMenu.add_command(label="Open current step recording...", command=self.onOpenCurrentStep)
        fileMenu.add_command(label="Open voltage step recording...", command=self.onOpenVoltageStep)
        fileMenu.add_command(label="Exit", command=self.onExit)    
        menubar.add_cascade(label="File", menu=fileMenu)     
        
        optionMenu = Menu(menubar, tearoff=0)
        optionMenu.add_command(label="Set parent file directory for recordings", command=self.askdirectory)
        optionMenu.add_command(label="Set directory for generated excel files", command=self.askdirectorySave)
        menubar.add_cascade(label="Options", menu=optionMenu)           

        optionMenu = Menu(menubar, tearoff=0)
        optionMenu.add_command(label="Read Basic SOP file", command=self.onOpenBasicSOP)
        optionMenu.add_command(label="Read synaptic SOP file", command=self.onOpenSynapticSOP)
        menubar.add_cascade(label="Analysis", menu=optionMenu)               
        
        helpMenu = Menu(menubar, tearoff=0)
        helpMenu.add_command(label="Help", command=self.onHelp)
        helpMenu.add_command(label="About", command=self.onAbout)    
        menubar.add_cascade(label="?", menu=helpMenu)              

    
        #self.txt = Text(self)
        #self.txt.pack(fill=BOTH, expand=1)
        
    def askdirectory(self):
        global openDirectory, saveDirectory, dest_filename, dest_directory
        """Returns a selected directoryname."""
        openDirectory = tkFileDialog.askdirectory()
        saveDirectory = openDirectory
        dest_directory = openDirectory
        print dest_directory
        
    def askdirectorySave(self):
        global saveDirectory, dest_filename, dest_directory
        """Returns a selected directoryname."""
        saveDirectory = tkFileDialog.askdirectory()  
        dest_directory  = saveDirectory
        #dest_filename = saveDirectory +  "\\"  + abffile[0] + ".xlsx"
        print dest_directory

    def onExit(self):
        root.destroy()
        
    def onAbout(self):
        box.showinfo("About Basic Properties Analyzer", "Version 1.2, February 2016\n\n Copyright: Christian Schnell (cschnell@schnell-thiessen.de)\n\n https://github.com/schrist81/BasicPropertyAnalyzer") 
    
    def onHelp(self):
        pass     
    
    def onOpenVoltageStep(self):
        global rec, complete_dataset, ws1
        ftypes = [('Axon binary files', '*.abf'), ('All files', '*')]
        dlg = tkFileDialog.Open(self, filetypes = ftypes, initialdir = openDirectory)
        fl = dlg.show()
        
        if fl != '':
            rec = stfio.read(str(fl))
            path = str(fl)
            
            #Extract filename for Excel file
            singles = path.split("/")
            iAP_file = singles[-1]      
        

            i = 0
            K_max_current = 0
            Na_act_max_current = 0
            # voltage step inserter for K mean
            voltageStepInserter(360, 396, -120, 5)
            # voltage step inserter for K current density
            voltageStepInserter(399, 435, -120, 5)        
            # voltage step inserter for Na activation currents
            voltageStepInserter(85, 121, -120, 5)         
            # voltage step inserter for Na activation current densities
            voltageStepInserter(123, 159, -120, 5)      
            # voltage step inserter for Na activation current densities driving force correction
            voltageStepInserter(162, 198, -120, 5)        
            # voltage step inserter for Na activation current densities driving force correction normalized conductance
            voltageStepInserter(202, 238, -120, 5)  
            
            # voltage step inserter for Na activation currents
            voltageStepInserter(242, 278, -120, 5)          
            # voltage step inserter for Na inactivation driving force correction
            voltageStepInserter(281, 317, -120, 5)         
            # voltage step inserter for Na inactivation driving force correction normalized conductance
            voltageStepInserter(321, 357, -120, 5)                
            
            while i < len(rec[0]):
                
                trace = rec[0][i].asarray()
                
                ''' Na activation currents 
                '''                
                # sampling rate: rec.dt in ms, mean of interval between 58 and 100 ms
                # look for negative peak later
                Na_activation_interval_begin = 58/rec.dt
                Na_activation_interval_end = 100/rec.dt
                
                # Na activation currents
                min_activation_peak = peakdetect(trace[Na_activation_interval_begin:Na_activation_interval_end], "negative", None, lookahead = 14 , delta=0)               
                #print str(i)+": "+str(min_activation_peak)
                coordinateNaAct = "B" + str(85+i)
                ws1[coordinateNaAct] = min_activation_peak         
                
                # Na activation current densities
                coordinate = "B" + str(123+i)
                field = "="+coordinateNaAct+"/B$9"
                ws1[coordinate] = field             
                
                # Na activation current densities - driving force correction
                coordinateNaActDriForce = "B" + str(162+i)
                ws1[coordinateNaActDriForce] = "="+str(coordinate)+"/($A"+str(162+i)+"-66.68)"
                
                # Na activation current densities - driving force correction - normalized conductance
                coordinateNaActDriForceConduct = "B" + str(202+i)
                ws1[coordinateNaActDriForceConduct] = "=B"+str(162+i)+"/MAX(B162:B198)"  
                
                # save maximum Na current
                if min_activation_peak < Na_act_max_current:
                    Na_act_max_current = min_activation_peak                
                Na_act_max_current_density = "="+str(Na_act_max_current)+"/B9"
                ws1['B81'] = Na_act_max_current_density
                
                
                ''' Na inactivation currents
                '''
                # sampling rate: rec.dt in ms, mean of interval between 50 and 100 ms
                # look for negative peak later
                Na_inactivation_interval_begin = 250/rec.dt
                Na_inactivation_interval_end = 270/rec.dt              
                
				# Na inactivation currents
                min_inactivation_peak = peakdetect(trace[Na_inactivation_interval_begin:Na_inactivation_interval_end], "negative", None, lookahead = 10 , delta=0)               
                #print str(i)+": "+str(min_activation_peak)
                coordinateNaInact = "B" + str(242+i)
                ws1[coordinateNaInact] = min_inactivation_peak         
                
                # Na inactivation driving force correction
                coordinateNaInactDriForce = "B" + str(281+i)
                ws1[coordinateNaInactDriForce] = "="+str(coordinateNaInact)+"/($A"+str(281+i)+"-66.68)"
                
                # Na inactivation - driving force correction - normalized conductance
                coordinateNaInactDriForceConduct = "B" + str(321+i)
                ws1[coordinateNaInactDriForceConduct] = "=B"+str(281+i)+"/MAX(B281:B317)"                  
                #=B281/MAX(B$281:B$317)
                ''' K currents 
                '''
                # sampling rate: rec.dt in ms, mean of interval between 208 and 258 ms
                K_mean_interval_begin = 208/rec.dt
                K_mean_interval_end = 258/rec.dt                    
                
                # determine mean K currents and fill in table
                coordinateA = "B" + str(360+i)
                ws1[coordinateA] = trace[K_mean_interval_begin:K_mean_interval_end].mean()
                # determine mean K current density and fill in table
                coordinate = "B" + str(399+i)
                field = "="+coordinateA+"/B9"
                ws1[coordinate] = field
                # save maximum potassium current
                mean = trace[K_mean_interval_begin:K_mean_interval_end].mean()
                if mean > K_max_current:
                    K_max_current = mean

                i = i + 1
                
        
            # save maximum K+ current
            K_max_current_density = "="+str(K_max_current)+"/B9"
            ws1['B82'] = K_max_current_density
            print "Inactivation protocol"
        wb.save(filename = dest_filename)
    
    def onOpenCurrentStep(self):
        global rec, complete_dataset, iAP_file, input_resistance, ws1
        ftypes = [('Axon binary files', '*.abf'), ('All files', '*')]
        dlg = tkFileDialog.Open(self, filetypes = ftypes, initialdir = openDirectory)
        fl = dlg.show()
        
        if fl != '':
            rec = stfio.read(str(fl))
            path = str(fl)
            
            #Extract filename for Excel file
            singles = path.split("/")
            iAP_file = singles[-1]      
        
            #Delete overshoot, AHP, sweep number and spike height in case
            #a file with no spikes is opened after one with spikes
            ws1['B27'] = ""            
            ws1['B28'] = ""
            ws1['B30'] = ""
            ws1['B31'] = ""
            ws1['B32'] = ""
            #set sweep number to -1 to set it later to the first sweep with an AP
            sweep = -1

            for i in range(0,len(rec[0])):
                #print str(i)+": "
                #trace = rec[0][i].asarray()
                trace = np.array(rec[0][i])  
                
                
                #sampling rate: rec.dt in ms, mean of interval between 0.5 and 1 s
                interval_begin = 500/rec.dt
                interval_end = 1000/rec.dt
            
                #extracts the 1s current step part of each sweep
                injected_trace = trace[1612:51612]    

                if (i == 0):
                    if (injected_trace[injected_trace.argmax()] < 0):
                        first_mean = trace[interval_begin:interval_end].mean()
                    else:
                        first_mean = "Error"
                elif (i == 1): 
                    if (injected_trace[injected_trace.argmax()] < 0):
                        second_mean = trace[interval_begin:interval_end].mean()
                    else:
                        second_mean = "Error"
                i = i + 1
             
                coordinate = "B" + str(36+i)
                if (injected_trace[injected_trace.argmax()] > 0):
                    if sweep == -1:
                        ws1['B28'] = i
                        sweep = i
                        # Determine Overshoot
                        overshoot = 0
                        afterhyperpolarization = 0
                        firstPeak = peakdetect(injected_trace, "positive", None, lookahead = 300, delta=0)[0]
                        ws1['B30'] = firstPeak[1]

                        # Determine Afterhyperpolarization
                        afterhyperpolarization_trace = injected_trace[firstPeak[0]:firstPeak[0]+(80/rec.dt)]
                        afterhyperpolarization = peakdetect(afterhyperpolarization_trace, "negative", None, lookahead = 300, delta=0) 

                        ws1['B31'] = afterhyperpolarization     
                        # Determine Spike Height
                        ws1['B32'] = "=B30-B31"
                        #ws1['B32'] = firstPeak[1] - afterhyperpolarization
                    ws1[coordinate] = len(peakdetect(injected_trace, "positive", None, lookahead = 300, delta=5))
                else:
                    ws1[coordinate] = 0
        
        ws1['B8'] = '' # leert input resistance, falls der Wert nicht korrekt berechnet werden kann, steht ansonsten noch der input resistance der letzten Zelle drin.
        ws1['B8'] = inducedActivity().calculateInputResistance(first_mean, second_mean)  
        ws1['B27'] = iAP_file
        print "Current step"
        wb.save(filename = dest_filename)
            
    
        
    def onOpenGapFree(self):
        completeList = []
        global rec, complete_dataset, abffile, dest_filename
        ftypes = [('Axon binary files', '*.abf'), ('All files', '*')]
        dlg = tkFileDialog.Open(self, filetypes = ftypes, initialdir = openDirectory)
        fl = dlg.show()
        
        if fl != '':
            rec = stfio.read(str(fl))
            path = str(fl)
            
            #Extract filename for Excel file
            singles = path.split("/")
            abffile = singles[-1]
            abffile = abffile.split(".")
            ws1['B3'] = singles[-3]+"/"+singles[-2]
            

            # Read sampling interval in ms -> 0.2 ms for testing file
            # 1 section mit 256 Datenpunkten entspricht 51 ms bei 0.2 sampling interval
            # print(rec.dt)
            
            # constructs single numpy array from recording
            for i in range(0,len(rec[0])):
                completeList.extend(rec[0][i])
            complete_dataset = np.array(completeList)
            
            
            # noch nicht ganz sauber, muss in separate Klasse
            ws1['B7'] = '' #Feld für RMP auf leer setzen, damit nicht der Wert der vorherigen Zelle eingetragen wird, falls Zelle sAP oder attempted sAP zeigt.
            if spontActivity().action_potential_found():
                sAP =  3
                ws1['B15'] = sAP
                ws1['B13'] = ''
                ws1['B14'] = ''

            else:
                if spontActivity().attempted_action_potential_found():
                    sAP =  2
                    ws1['B14'] = sAP
                    ws1['B13'] = ''
                    ws1['B15'] = ''
                else:
                    sAP =  1
                    ws1['B13'] = sAP
                    ws1['B14'] = ''
                    ws1['B15'] = ''
                    #RMP only determined in sAP code 1 files (other cases more complicated, maybe later)
                    ws1['B7'] = spontActivity().mean_first_10s(complete_dataset)   

            ws1['B16'] = "=SUM(B13:B15)"
            ws1['B6'] = rig
            ws1['B5'] = abffile[0] + "." + abffile[1]
            dest_filename = dest_directory + "\\"  + abffile[0] + ".xlsx"
            wb.save(filename = dest_filename)
            print "Gap free"
    def readFile(self, filename):
        pass
        #f = open(filename, "r")
        #text = f.read()
        #return text

    def onOpenBasicSOP(self):
        ftypes = [('Excel files', '*.xlsx'), ('All files', '*')]
        dlg = tkFileDialog.Open(self, filetypes = ftypes, initialdir = openDirectory)
        fl = dlg.show()
        BasicSOPworkbook = openpyxl.load_workbook(fl)
        BasicSOPsheets = BasicSOPworkbook.get_sheet_names()
        RMPlistoflists = []
        inputResistanceListofLists = []
        capacitanceListofLists = []
        thresholdListofLists = []
        overshootListofLists = []
        afterhyperpolarizationListofLists = []
        spikeHeightListofLists = []
        depolRateListofLists = []
        repolRateListofLists = []
        halfWidthListofLists = []    
        maxNaListofLists = []
        maxKListofLists = []      
        for i in range(0,len(BasicSOPsheets)):

            # #############################################
            # plot figures of spontaneous action potentials
            # #############################################
            sAP_silent = 0
            numberOfCells = 0
            sheet = BasicSOPworkbook.get_sheet_by_name(BasicSOPsheets[i])
            for rowOfCellObjects in tuple(sheet['B13':str(get_column_letter(sheet.get_highest_column()))+str(13)]):
                for cellObj in rowOfCellObjects:
                    if cellObj.value == 1:
                        sAP_silent = sAP_silent + cellObj.value

            numberOfCells = sAP_silent * 1.0
        
            sAP_attempted = 0
            sheet = BasicSOPworkbook.get_sheet_by_name(BasicSOPsheets[i])
            for rowOfCellObjects in tuple(sheet['B14':str(get_column_letter(sheet.get_highest_column()))+str(14)]):
                for cellObj in rowOfCellObjects:
                    if cellObj.value == 2:
                        sAP_attempted = sAP_attempted + cellObj.value
            sAP_attempted = sAP_attempted/2.0
            numberOfCells = numberOfCells + sAP_attempted

            sAP_full = 0
            sheet = BasicSOPworkbook.get_sheet_by_name(BasicSOPsheets[i])
            for rowOfCellObjects in tuple(sheet['B15':str(get_column_letter(sheet.get_highest_column()))+str(15)]):
                for cellObj in rowOfCellObjects:
                    if cellObj.value == 3.0:
                        sAP_full = sAP_full + cellObj.value
            sAP_full = sAP_full/3            
            numberOfCells = numberOfCells + sAP_full
            
            sizes = [(sAP_full/numberOfCells)*100*360, (sAP_attempted/numberOfCells)*100*360, (sAP_silent/numberOfCells)*100*360]
            colors = ['green','yellow','red']
            #explode = (0, 0, 0)  # explode 1st slice
                        
            plt.yticks
            plt.suptitle(BasicSOPsheets[i], fontsize=20)
            plt.pie(sizes, colors=colors,
                    autopct='%1.1f%%', shadow=False, startangle=90)
            plt.axis('equal')
            plt.savefig('sAP'+str(i), dpi=None, facecolor='w', edgecolor='w',
                    orientation='portrait', papertype=None, format=None,
                    transparent=False, bbox_inches=None, pad_inches=0.1,
                    frameon=None)
            plt.clf()

            sAP_silent = 0
            numberOfCells = 0
            sheet = BasicSOPworkbook.get_sheet_by_name(BasicSOPsheets[i])
            for rowOfCellObjects in tuple(sheet['B13':str(get_column_letter(sheet.get_highest_column()))+str(13)]):
                for cellObj in rowOfCellObjects:
                    if cellObj.value == 1:
                        sAP_silent = sAP_silent + cellObj.value

            numberOfCells = sAP_silent * 1.0
            
            # #########################################
            # plot figures of induced action potentials
            # #########################################
            iAP_none = 0
            sheet = BasicSOPworkbook.get_sheet_by_name(BasicSOPsheets[i])
            for rowOfCellObjects in tuple(sheet['B19':str(get_column_letter(sheet.get_highest_column()))+str(19)]):
                for cellObj in rowOfCellObjects:
                    if cellObj.value == 1:
                        iAP_none = iAP_none + cellObj.value
            iAP_none = iAP_none/1.0
            numberOfCells = numberOfCells + iAP_none

            iAP_single_attempted = 0
            sheet = BasicSOPworkbook.get_sheet_by_name(BasicSOPsheets[i])
            for rowOfCellObjects in tuple(sheet['B20':str(get_column_letter(sheet.get_highest_column()))+str(20)]):
                for cellObj in rowOfCellObjects:
                    if cellObj.value == 2:
                        iAP_single_attempted = iAP_single_attempted + cellObj.value
            iAP_single_attempted = iAP_single_attempted/2.0
            numberOfCells = numberOfCells + iAP_single_attempted


            iAP_single = 0
            sheet = BasicSOPworkbook.get_sheet_by_name(BasicSOPsheets[i])
            for rowOfCellObjects in tuple(sheet['B21':str(get_column_letter(sheet.get_highest_column()))+str(21)]):
                for cellObj in rowOfCellObjects:
                    if cellObj.value == 3:
                        iAP_single = iAP_single + cellObj.value
            iAP_single = iAP_single/3.0
            numberOfCells = numberOfCells + iAP_single

            iAP_train_attempted = 0
            sheet = BasicSOPworkbook.get_sheet_by_name(BasicSOPsheets[i])
            for rowOfCellObjects in tuple(sheet['B22':str(get_column_letter(sheet.get_highest_column()))+str(22)]):
                for cellObj in rowOfCellObjects:
                    if cellObj.value == 4:
                        iAP_train_attempted = iAP_train_attempted + cellObj.value
            iAP_train_attempted = iAP_train_attempted/4.0
            numberOfCells = numberOfCells + iAP_train_attempted

            iAP_train = 0
            sheet = BasicSOPworkbook.get_sheet_by_name(BasicSOPsheets[i])
            for rowOfCellObjects in tuple(sheet['B23':str(get_column_letter(sheet.get_highest_column()))+str(23)]):
                for cellObj in rowOfCellObjects:
                    if cellObj.value == 5:
                        iAP_train = iAP_train + cellObj.value
            iAP_train = iAP_train/5.0
            numberOfCells = numberOfCells + iAP_train               
            
            sizes = [(iAP_train/numberOfCells)*100*360, (iAP_train_attempted/numberOfCells)*100*360, (iAP_single/numberOfCells)*100*360, (iAP_single_attempted/numberOfCells)*100*360, (iAP_none/numberOfCells)*100*360]
            colors = ['green','lightgreen','yellow','orange','red']
            #explode = (0, 0, 0)  # explode 1st slice
            # Plot
            
            plt.yticks
            plt.suptitle(BasicSOPsheets[i], fontsize=20)
            plt.pie(sizes, colors=colors,
                    autopct='%1.1f%%', shadow=False, startangle=90)
            plt.axis('equal')
            plt.savefig('iAP'+str(i), dpi=None, facecolor='w', edgecolor='w',
                    orientation='portrait', papertype=None, format=None,
                    transparent=False, bbox_inches=None, pad_inches=0.1,
                    frameon=None)
            plt.clf()


            # #########################################
            # plot of basic properties
            # #########################################
            RMPlist = []
            for rowOfCellObjects in tuple(sheet['B7':str(get_column_letter(sheet.get_highest_column()))+str(7)]):
                for cellObj in rowOfCellObjects:
                    if type(cellObj.value) == float:
                        RMPlist.append(cellObj.value)
                RMPlistoflists.append(RMPlist)

            inputResistanceList = []
            for rowOfCellObjects in tuple(sheet['B8':str(get_column_letter(sheet.get_highest_column()))+str(8)]):
                for cellObj in rowOfCellObjects:
                    if type(cellObj.value) == float:
                        inputResistanceList.append(cellObj.value)
                inputResistanceListofLists.append(inputResistanceList)

            capacitanceList = []
            for rowOfCellObjects in tuple(sheet['B9':str(get_column_letter(sheet.get_highest_column()))+str(9)]):
                for cellObj in rowOfCellObjects:
                    if type(cellObj.value) == float:
                        capacitanceList.append(cellObj.value)
                capacitanceListofLists.append(capacitanceList)              

            thresholdList = []
            for rowOfCellObjects in tuple(sheet['B29':str(get_column_letter(sheet.get_highest_column()))+str(29)]):
                for cellObj in rowOfCellObjects:
                    if type(cellObj.value) == float:
                        thresholdList.append(cellObj.value)
                thresholdListofLists.append(thresholdList)    

            overshootList = []
            for rowOfCellObjects in tuple(sheet['B30':str(get_column_letter(sheet.get_highest_column()))+str(30)]):
                for cellObj in rowOfCellObjects:
                    if type(cellObj.value) == float:
                        overshootList.append(cellObj.value)
                overshootListofLists.append(overshootList)  

            afterhypolarizationList = []
            for rowOfCellObjects in tuple(sheet['B31':str(get_column_letter(sheet.get_highest_column()))+str(31)]):
                for cellObj in rowOfCellObjects:
                    if type(cellObj.value) == float:
                        afterhypolarizationList.append(cellObj.value)
                afterhyperpolarizationListofLists.append(afterhypolarizationList)  

            spikeHeightList = []
            for rowOfCellObjects in tuple(sheet['B32':str(get_column_letter(sheet.get_highest_column()))+str(32)]):
                for cellObj in rowOfCellObjects:
                    if type(cellObj.value) == float:
                        spikeHeightList.append(cellObj.value)
                spikeHeightListofLists.append(spikeHeightList)   

            depolRateList = []
            for rowOfCellObjects in tuple(sheet['B33':str(get_column_letter(sheet.get_highest_column()))+str(33)]):
                for cellObj in rowOfCellObjects:
                    if type(cellObj.value) == float:
                        depolRateList.append(cellObj.value)
                depolRateListofLists.append(depolRateList)            


            repolRateList = []
            for rowOfCellObjects in tuple(sheet['B34':str(get_column_letter(sheet.get_highest_column()))+str(34)]):
                for cellObj in rowOfCellObjects:
                    if type(cellObj.value) == float:
                        repolRateList.append(cellObj.value)
                repolRateListofLists.append(repolRateList)      

            halfWidthList = []
            for rowOfCellObjects in tuple(sheet['B35':str(get_column_letter(sheet.get_highest_column()))+str(35)]):
                for cellObj in rowOfCellObjects:
                    if type(cellObj.value) == float:
                        halfWidthList.append(cellObj.value)
                halfWidthListofLists.append(halfWidthList)            

            maxNaList = []
            for rowOfCellObjects in tuple(sheet['B81':str(get_column_letter(sheet.get_highest_column()))+str(81)]):
                for cellObj in rowOfCellObjects:
                    if type(cellObj.value) == float:
                        maxNaList.append(cellObj.value)
                maxNaListofLists.append(maxNaList)

            maxKList = []
            for rowOfCellObjects in tuple(sheet['B82':str(get_column_letter(sheet.get_highest_column()))+str(82)]):
                for cellObj in rowOfCellObjects:
                    if type(cellObj.value) == float:
                        maxKList.append(cellObj.value)
                maxKListofLists.append(maxKList)



        #hier müssen jetzt die nparrays der einzelnen Datenblätter in ein nparray pro Parameter gespeichert werden, um sie plotten und analysieren zu können
        colors = ['#000000', '#666666', '#999999', '#e6e6e6', '#339933', '#66ff66', '#0033cc', '#3399ff', '#991f00', '#ff3300', '#ff9980', '#000000', '#666666', '#999999', '#e6e6e6', '#339933', '#0033cc', '#66ff66', '#3399ff', '#991f00', '#ff3300', '#ff9980']
        RMPnparray=np.array([np.array(xi) for xi in RMPlistoflists])
        listForPlotting = []
        listForError = []
        for n in range(len(RMPnparray)):
            listForPlotting.append(RMPnparray[n].mean())
            listForError.append(stats.sem(RMPnparray[n]))
        x = np.arange(len(RMPnparray))
        #plt.suptitle("Resting membrane potential", fontsize=14)
        plt.xticks(rotation=50, horizontalalignment='right')
        
        plt.bar(x, listForPlotting, yerr=listForError, ecolor='#000000', align='center', color=colors)

        #Following 5 lines of code are necessary to remove the right and top spine of the bar graph
        ax = plt.subplot(111)
        # Hide the right and top spines
        ax.spines['right'].set_visible(False)
        ax.spines['bottom'].set_visible(False)
        # Only show ticks on the left and bottom spines
        ax.yaxis.set_ticks_position('left')
        ax.xaxis.set_ticks_position('top')
        
        plt.xticks(x, BasicSOPsheets)
        plt.ylabel('Resting membrane potential [mV]', fontsize=20, x=0.05, style='italic')
        plt.autoscale()
        plt.margins(0.05, 0)
        plt.tight_layout()

        plt.savefig('RMP', dpi=None, facecolor='w', edgecolor='w',
                orientation='portrait', papertype=None, format=None,
                transparent=False, bbox_inches=None, pad_inches=0.1,
                frameon=None)
        plt.clf()      

        IPRnparray=np.array([np.array(xi) for xi in inputResistanceListofLists])
        listForPlotting = []
        listForError = []
        for n in range(len(IPRnparray)):
            listForPlotting.append(IPRnparray[n].mean())
            listForError.append(stats.sem(IPRnparray[n]))
        x = np.arange(len(IPRnparray))
        print len(IPRnparray)
        #x = np.arange(len(IPRnparray))

        plt.suptitle("Input resistance", y=1.05, fontsize=14)
        plt.xticks(rotation=50, horizontalalignment='right')
        plt.bar(x, listForPlotting, yerr=listForError, ecolor='#000000', align='center', color=colors)

        #Following 5 lines of code are necessary to remove the right and top spine of the bar graph
        ax = plt.subplot(111)
        # Hide the right and top spines
        ax.spines['right'].set_visible(False)
        ax.spines['top'].set_visible(False)
        # Only show ticks on the left and bottom spines
        ax.yaxis.set_ticks_position('left')
        ax.xaxis.set_ticks_position('bottom')
        
        plt.xticks(x, BasicSOPsheets)
        plt.ylabel('Input resistance [MOhm]', fontsize=20, x=0.05, style='italic')
        plt.autoscale()
        plt.margins(0.05, 0)
        plt.tight_layout()
        
        plt.savefig('InputResistance', dpi=None, facecolor='w', edgecolor='w',
                orientation='landscape', papertype=None, format=None,
                transparent=False, bbox_inches=None, pad_inches=0.1,
                frameon=None)
        plt.clf()     


        capacitancenparray=np.array([np.array(xi) for xi in capacitanceListofLists])
        listForPlotting = []
        listForError = []
        for n in range(len(capacitancenparray)):
            listForPlotting.append(capacitancenparray[n].mean())
            listForError.append(stats.sem(capacitancenparray[n]))
        x = np.arange(len(capacitancenparray))

        #x = np.arange(len(IPRnparray))

        plt.suptitle("Capacitance", y=1.05, fontsize=14)
        plt.xticks(rotation=50, horizontalalignment='right')
        plt.bar(x, listForPlotting, yerr=listForError, ecolor='#000000', align='center', color=colors)

        #Following 5 lines of code are necessary to remove the right and top spine of the bar graph
        ax = plt.subplot(111)
        # Hide the right and top spines
        ax.spines['right'].set_visible(False)
        ax.spines['top'].set_visible(False)
        # Only show ticks on the left and bottom spines
        ax.yaxis.set_ticks_position('left')
        ax.xaxis.set_ticks_position('bottom')
        
        plt.xticks(x, BasicSOPsheets)
        plt.ylabel('Capacitance [pF]', fontsize=20, x=0.05, style='italic')
        plt.autoscale()
        plt.margins(0.05, 0)
        plt.tight_layout()
        
        plt.savefig('Capacitance', dpi=None, facecolor='w', edgecolor='w',
                orientation='landscape', papertype=None, format=None,
                transparent=False, bbox_inches=None, pad_inches=0.1,
                frameon=None)
        plt.clf() 



        thresholdnparray=np.array([np.array(xi) for xi in thresholdListofLists])
        listForPlotting = []
        listForError = []
        for n in range(len(thresholdnparray)):
            listForPlotting.append(thresholdnparray[n].mean())
            listForError.append(stats.sem(thresholdnparray[n]))
        x = np.arange(len(thresholdnparray))

        #x = np.arange(len(IPRnparray))

        #plt.suptitle("Capacitance", y=1.05, fontsize=14)
        plt.xticks(rotation=50, horizontalalignment='right')
        plt.bar(x, listForPlotting, yerr=listForError, ecolor='#000000', align='center', color=colors)

        #Following 5 lines of code are necessary to remove the right and top spine of the bar graph
        ax = plt.subplot(111)
        # Hide the right and top spines
        ax.spines['right'].set_visible(False)
        ax.spines['bottom'].set_visible(False)
        # Only show ticks on the left and bottom spines
        ax.yaxis.set_ticks_position('left')
        ax.xaxis.set_ticks_position('bottom')
        
        plt.xticks(x, BasicSOPsheets)
        plt.ylabel('Threshold [mV]', fontsize=20, x=0.05, style='italic')
        plt.autoscale()
        plt.margins(0.05, 0)
        plt.tight_layout()
        
        plt.savefig('Threshold', dpi=None, facecolor='w', edgecolor='w',
                orientation='landscape', papertype=None, format=None,
                transparent=False, bbox_inches=None, pad_inches=0.1,
                frameon=None)
        plt.clf() 


        overshootnparray=np.array([np.array(xi) for xi in overshootListofLists])
        listForPlotting = []
        listForError = []
        for n in range(len(overshootnparray)):
            listForPlotting.append(overshootnparray[n].mean())
            listForError.append(stats.sem(overshootnparray[n]))
        x = np.arange(len(overshootnparray))

        #x = np.arange(len(IPRnparray))

        #plt.suptitle("Capacitance", y=1.05, fontsize=14)
        plt.xticks(rotation=50, horizontalalignment='right')
        plt.bar(x, listForPlotting, yerr=listForError, ecolor='#000000', align='center', color=colors)

        #Following 5 lines of code are necessary to remove the right and top spine of the bar graph
        ax = plt.subplot(111)
        # Hide the right and top spines
        ax.spines['right'].set_visible(False)
        ax.spines['top'].set_visible(False)
        # Only show ticks on the left and bottom spines
        ax.yaxis.set_ticks_position('left')
        ax.xaxis.set_ticks_position('bottom')
        
        plt.xticks(x, BasicSOPsheets)
        plt.ylabel('Overshoot [mV]', fontsize=20, x=0.05, style='italic')
        plt.autoscale()
        plt.margins(0.05, 0)
        plt.tight_layout()
        
        plt.savefig('Overshoot', dpi=None, facecolor='w', edgecolor='w',
                orientation='landscape', papertype=None, format=None,
                transparent=False, bbox_inches=None, pad_inches=0.1,
                frameon=None)
        plt.clf()


       

        afterhypolarizationnparray=np.array([np.array(xi) for xi in afterhyperpolarizationListofLists])
        listForPlotting = []
        listForError = []
        for n in range(len(afterhypolarizationnparray)):
            listForPlotting.append(afterhypolarizationnparray[n].mean())
            listForError.append(stats.sem(afterhypolarizationnparray[n]))
        x = np.arange(len(afterhypolarizationnparray))

        #x = np.arange(len(IPRnparray))

        #plt.suptitle("Spike Height", y=1.05, fontsize=14)
        plt.xticks(rotation=50, horizontalalignment='right')
        plt.bar(x, listForPlotting, yerr=listForError, ecolor='#000000', align='center', color=colors)

        #Following 5 lines of code are necessary to remove the right and top spine of the bar graph
        ax = plt.subplot(111)
        # Hide the right and top spines
        ax.spines['right'].set_visible(False)
        ax.spines['bottom'].set_visible(False)
        # Only show ticks on the left and bottom spines
        ax.yaxis.set_ticks_position('left')
        ax.xaxis.set_ticks_position('top')
        
        plt.xticks(x, BasicSOPsheets)
        plt.ylabel('Afterhyperpolarization [mV]', fontsize=20, x=0.05, style='italic')
        plt.autoscale()
        plt.margins(0.05, 0)
        plt.tight_layout()
        
        plt.savefig('Afterhypolarization', dpi=None, facecolor='w', edgecolor='w',
                orientation='landscape', papertype=None, format=None,
                transparent=False, bbox_inches=None, pad_inches=0.1,
                frameon=None)
        plt.clf()         


        depolRatenparray=np.array([np.array(xi) for xi in depolRateListofLists])
        listForPlotting = []
        listForError = []
        for n in range(len(depolRatenparray)):
            listForPlotting.append(depolRatenparray[n].mean())
            listForError.append(stats.sem(depolRatenparray[n]))
        x = np.arange(len(depolRatenparray))

        #x = np.arange(len(IPRnparray))

        #plt.suptitle("Capacitance", y=1.05, fontsize=14)
        plt.xticks(rotation=50, horizontalalignment='right')
        plt.bar(x, listForPlotting, yerr=listForError, ecolor='#000000', align='center', color=colors)

        #Following 5 lines of code are necessary to remove the right and top spine of the bar graph
        ax = plt.subplot(111)
        # Hide the right and top spines
        ax.spines['right'].set_visible(False)
        ax.spines['top'].set_visible(False)
        # Only show ticks on the left and bottom spines
        ax.yaxis.set_ticks_position('left')
        ax.xaxis.set_ticks_position('bottom')
        
        plt.xticks(x, BasicSOPsheets)
        plt.ylabel('Maximal depolarization rate [mV/ms]', fontsize=17.5, x=0.05, style='italic')
        plt.autoscale()
        plt.margins(0.05, 0)
        plt.tight_layout()
        
        plt.savefig('Depolarization', dpi=None, facecolor='w', edgecolor='w',
                orientation='landscape', papertype=None, format=None,
                transparent=False, bbox_inches=None, pad_inches=0.1,
                frameon=None)
        plt.clf()

        repolratenparray=np.array([np.array(xi) for xi in repolRateListofLists])
        listForPlotting = []
        listForError = []
        for n in range(len(repolratenparray)):
            listForPlotting.append(repolratenparray[n].mean())
            listForError.append(stats.sem(repolratenparray[n]))
        x = np.arange(len(thresholdnparray))

        #x = np.arange(len(IPRnparray))

        #plt.suptitle("Capacitance", y=1.05, fontsize=14)
        plt.xticks(rotation=50, horizontalalignment='right')
        plt.bar(x, listForPlotting, yerr=listForError, ecolor='#000000', align='center', color=colors)

        #Following 5 lines of code are necessary to remove the right and top spine of the bar graph
        ax = plt.subplot(111)
        # Hide the right and top spines
        ax.spines['right'].set_visible(False)
        ax.spines['bottom'].set_visible(False)
        # Only show ticks on the left and bottom spines
        ax.yaxis.set_ticks_position('left')
        ax.xaxis.set_ticks_position('top')
        
        plt.xticks(x, BasicSOPsheets)
        plt.ylabel('Maximal repolarization rate [mV/ms]', fontsize=18, x=0.05, style='italic')
        plt.autoscale()
        plt.margins(0.05, 0)
        plt.tight_layout()
        
        plt.savefig('Repolarization', dpi=None, facecolor='w', edgecolor='w',
                orientation='landscape', papertype=None, format=None,
                transparent=False, bbox_inches=None, pad_inches=0.1,
                frameon=None)
        plt.clf() 

        halfWidthRatenparray=np.array([np.array(xi) for xi in halfWidthListofLists])
        listForPlotting = []
        listForError = []
        for n in range(len(halfWidthRatenparray)):
            listForPlotting.append(halfWidthRatenparray[n].mean())
            listForError.append(stats.sem(halfWidthRatenparray[n]))
        x = np.arange(len(halfWidthRatenparray))

        #x = np.arange(len(IPRnparray))

        #plt.suptitle("Capacitance", y=1.05, fontsize=14)
        plt.xticks(rotation=50, horizontalalignment='right')
        plt.bar(x, listForPlotting, yerr=listForError, ecolor='#000000', align='center', color=colors)

        #Following 5 lines of code are necessary to remove the right and top spine of the bar graph
        ax = plt.subplot(111)
        # Hide the right and top spines
        ax.spines['right'].set_visible(False)
        ax.spines['top'].set_visible(False)
        # Only show ticks on the left and bottom spines
        ax.yaxis.set_ticks_position('left')
        ax.xaxis.set_ticks_position('bottom')
        
        plt.xticks(x, BasicSOPsheets)
        plt.ylabel('Half width duration [ms]', fontsize=20, x=0.05, style='italic')
        plt.autoscale()
        plt.margins(0.05, 0)
        plt.tight_layout()
        
        plt.savefig('Halfwidth', dpi=None, facecolor='w', edgecolor='w',
                orientation='landscape', papertype=None, format=None,
                transparent=False, bbox_inches=None, pad_inches=0.1,
                frameon=None)
        plt.clf()

        maxKnparray=np.array([np.array(xi) for xi in maxKListofLists])
        listForPlotting = []
        listForError = []
        for n in range(len(maxKnparray)):
            listForPlotting.append(maxKnparray[n].mean())
            listForError.append(stats.sem(maxKnparray[n]))
        x = np.arange(len(maxKnparray))

        #x = np.arange(len(IPRnparray))

        #plt.suptitle("Capacitance", y=1.05, fontsize=14)
        plt.xticks(rotation=50, horizontalalignment='right')
        plt.bar(x, listForPlotting, yerr=listForError, ecolor='#000000', align='center', color=colors)

        #Following 5 lines of code are necessary to remove the right and top spine of the bar graph
        ax = plt.subplot(111)
        # Hide the right and top spines
        ax.spines['right'].set_visible(False)
        ax.spines['top'].set_visible(False)
        # Only show ticks on the left and bottom spines
        ax.yaxis.set_ticks_position('left')
        ax.xaxis.set_ticks_position('bottom')
        
        plt.xticks(x, BasicSOPsheets)
        plt.ylabel('Maximal K+ current density [pA/pF]', fontsize=20, x=0.05, style='italic')
        plt.autoscale()
        plt.margins(0.05, 0)
        plt.tight_layout()
        
        plt.savefig('maxK', dpi=None, facecolor='w', edgecolor='w',
                orientation='landscape', papertype=None, format=None,
                transparent=False, bbox_inches=None, pad_inches=0.1,
                frameon=None)
        plt.clf()        

        maxNanparray=np.array([np.array(xi) for xi in maxNaListofLists])
        listForPlotting = []
        listForError = []
        for n in range(len(maxNanparray)):
            listForPlotting.append(maxNanparray[n].mean())
            listForError.append(stats.sem(maxNanparray[n]))
        x = np.arange(len(maxNanparray))

        #x = np.arange(len(IPRnparray))

        #plt.suptitle("Capacitance", y=1.05, fontsize=14)
        plt.xticks(rotation=50, horizontalalignment='right')
        plt.bar(x, listForPlotting, yerr=listForError, ecolor='#000000', align='center', color=colors)

        #Following 5 lines of code are necessary to remove the right and top spine of the bar graph
        ax = plt.subplot(111)
        # Hide the right and top spines
        ax.spines['right'].set_visible(False)
        ax.spines['bottom'].set_visible(False)
        # Only show ticks on the left and bottom spines
        ax.yaxis.set_ticks_position('left')
        ax.xaxis.set_ticks_position('top')
        
        plt.xticks(x, BasicSOPsheets)
        plt.ylabel('Maximal Na+ current density [pA/pF]', fontsize=20, x=0.05, style='italic')
        plt.autoscale()
        plt.margins(0.05, 0)
        plt.tight_layout()
        
        plt.savefig('maxNa', dpi=None, facecolor='w', edgecolor='w',
                orientation='landscape', papertype=None, format=None,
                transparent=False, bbox_inches=None, pad_inches=0.1,
                frameon=None)
        plt.clf()  
 

    def onOpenSynapticSOP(self):
        pass
        
  

root = Tk()
ex = Example(root)
rig = tkSimpleDialog.askstring("Rig number", "Enter the number of your rig.")


root.geometry("300x250+300+300")
root.mainloop()
